from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import re
import os
import pandas as pd
class ExcelGenerator():
    def __init__(self, folder_path):
        self.folder_path = folder_path

    def export_to_xlsx(self, output_file):
        """
        Экспорт списка файлов и дубликатов в xlsx с добавлением расширения и размера файла.
        """
        def human_readable_size(size):
            """Форматирует размер файла в КБ, МБ и т.д."""
            for unit in ['Б', 'КБ', 'МБ', 'ГБ', 'ТБ']:
                if size < 1024.0:
                    return f"{size:.1f} {unit}"
                size /= 1024.0
            return f"{size:.1f} ТБ"

        file_names = self.get_file_names()

        if not file_names:
            print("Нет файлов для экспорта.")
            return

        duplicates = self.find_duplicates(file_names)
        self.duplicate_counter = len(duplicates)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "File List"

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Заголовки колонок
        ws.append(["Название файла", "Расширение", "Размер файла"])

        # Добавляем информацию о каждом файле
        for file_name in file_names:
            # Извлечение расширения и размера файла
            extension = os.path.splitext(file_name)[1].lower()  # Расширение файла
            file_size = os.path.getsize(os.path.join(self.folder_path, file_name))  # Размер файла
            readable_size = human_readable_size(file_size)  # Читаемый размер файла
            
            # Добавляем информацию в строку
            ws.append([file_name, extension, readable_size])
            
            # Выделяем дубликаты красным
            if file_name in duplicates:
                for col in range(1, 4):  # Первая, вторая и третья колонки
                    ws.cell(row=ws.max_row, column=col).fill = red_fill

        # Сохраняем Excel файл
        wb.save(output_file)


        
        for file_name in file_names:
            cell = ws.append([file_name])
            if file_name in duplicates:
                ws.cell(row=ws.max_row, column=1).fill = red_fill  # Выделяем ячейку красным

       
        wb.save(output_file)


    def find_duplicates(self, file_names):
        self.duplicate_counter = 0
        duplicates = []
        seen = {}

        for file in file_names:
            clean_name = self.clean_file_name(file)

            
            if clean_name in seen:
                duplicates.append(file)
            else:
                seen[clean_name] = file

        return duplicates
    
    def clean_file_name(self, file_name):
  
        name_without_extension, extension = os.path.splitext(file_name)     
        clean_name = re.sub(r'\s*\(\d+\)$', '', name_without_extension)

        
        return clean_name + extension
    
    def count_files(self):
        
        if not os.path.isdir(self.folder_path):
            print(f"Папка {self.folder_path} не существует")
            return 0
        file_count = sum([len(files) for _, _, files in os.walk(self.folder_path)])
        return file_count
    def get_file_names(self):
       
        if not os.path.isdir(self.folder_path):
            print(f"Папка {self.folder_path} не существует")
            return []
        file_names = []
        for _, _, files in os.walk(self.folder_path):
            file_names.extend(files) 

        return file_names
    
    def generate_hierarchy_report(self, root_folder, excel_path="Отчет по файлам.xlsx"):
        """
        Генерирует сводную таблицу в Excel для первых четырех уровней иерархии с информацией о расширениях файлов.
        Также включает папки-исключения "ОССЗ - Письма" и "ОССЗ - Письма проектировщика" как второй класс с указанным главным классом.
        """
       
        hierarchy_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int)))))

        for root, dirs, files in os.walk(root_folder):
            rel_path = os.path.relpath(root, root_folder).split(os.sep)

            if len(rel_path) >= 4:
                main_class, first_class, second_class, third_class = rel_path[:4]
               
                for file_name in files:
                    extension = os.path.splitext(file_name)[1][1:].lower()
                    hierarchy_data[main_class][first_class][second_class][third_class][extension] += 1
        

            elif len(rel_path) >= 2:
                main_class, potential_folder = rel_path[:2]
                if "ОССЗ - Письма"in potential_folder or "ОССЗ - Письма проектировщика" in potential_folder:
                    second_class = potential_folder
                    for file_name in files:
                        extension = os.path.splitext(file_name)[1][1:].lower()
                        hierarchy_data[main_class][None][second_class][None][extension] += 1
                elif "Прочие документы" in main_class:
             
                    for file_name in files:
                        extension = os.path.splitext(file_name)[1][1:].lower()
                        hierarchy_data[main_class][None][None][None][extension] += 1
     
        report_data = []

        for main_class, first_classes in hierarchy_data.items():
            for first_class, second_classes in first_classes.items():
                for second_class, third_classes in second_classes.items():
                    for third_class, extensions in third_classes.items():
                        row = {
                            "Проект": main_class,
                            "Классификатор код 1": first_class or "",
                            "Классификатор код 2": second_class,
                            "Классификатор код 3": third_class or ""
                        }
                        for ext, count in extensions.items():
                            row[ext] = count
                        report_data.append(row)


        df = pd.DataFrame(report_data)
        df.fillna(0, inplace=True)  

        with pd.ExcelWriter(excel_path) as writer:
            df.to_excel(writer, index=False, sheet_name="Отчет")