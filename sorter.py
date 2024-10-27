import os
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill
import re
import shutil
class Sorter():
    def __init__(self, folder_path):
        self.folder_path = folder_path

    def count_files(self):
        
        if not os.path.isdir(self.folder_path):
            print(f"Папка {self.folder_path} не существует")
            return 0
        file_count = sum([len(files) for _, _, files in os.walk(self.folder_path)])
        return file_count
    def get_file_names(self):
       
        if not os.path.isdir(self.folder_path):
            print(f"Папка {self.folder_path} не существует")
            return []
        file_names = []
        for _, _, files in os.walk(self.folder_path):
            file_names.extend(files) 

        return file_names
    def get_file_names(self):
        if not os.path.isdir(self.folder_path):
            print(f"Папка {self.folder_path} не существует")
            return []
        file_names = []
        for _, _, files in os.walk(self.folder_path):
            file_names.extend(files) 

        return file_names

    def clean_file_name(self, file_name):
  
        name_without_extension, extension = os.path.splitext(file_name)     
        clean_name = re.sub(r'\s*\(\d+\)$', '', name_without_extension)

        
        return clean_name + extension

    def find_duplicates(self, file_names):
        self.duplicate_counter = 0
        duplicates = []
        seen = {}

        for file in file_names:
            clean_name = self.clean_file_name(file)

            
            if clean_name in seen:
                duplicates.append(file)
            else:
                seen[clean_name] = file

        return duplicates
    
    

    def export_to_xlsx(self, output_file):
        """
        Экспорт списка файлов и дубликатов в xlsx с добавлением расширения и размера файла.
        """
        def human_readable_size(size):
            """Форматирует размер файла в КБ, МБ и т.д."""
            for unit in ['Б', 'КБ', 'МБ', 'ГБ', 'ТБ']:
                if size < 1024.0:
                    return f"{size:.1f} {unit}"
                size /= 1024.0
            return f"{size:.1f} ТБ"

        file_names = self.get_file_names()

        if not file_names:
            print("Нет файлов для экспорта.")
            return

        duplicates = self.find_duplicates(file_names)
        self.duplicate_counter = len(duplicates)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "File List"

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Заголовки колонок
        ws.append(["Название файла", "Расширение", "Размер файла"])

        # Добавляем информацию о каждом файле
        for file_name in file_names:
            # Извлечение расширения и размера файла
            extension = os.path.splitext(file_name)[1].lower()  # Расширение файла
            file_size = os.path.getsize(os.path.join(self.folder_path, file_name))  # Размер файла
            readable_size = human_readable_size(file_size)  # Читаемый размер файла
            
            # Добавляем информацию в строку
            ws.append([file_name, extension, readable_size])
            
            # Выделяем дубликаты красным
            if file_name in duplicates:
                for col in range(1, 4):  # Первая, вторая и третья колонки
                    ws.cell(row=ws.max_row, column=col).fill = red_fill

        # Сохраняем Excel файл
        wb.save(output_file)


        
        for file_name in file_names:
            cell = ws.append([file_name])
            if file_name in duplicates:
                ws.cell(row=ws.max_row, column=1).fill = red_fill  # Выделяем ячейку красным

       
        wb.save(output_file)


    def replace_cyrillic_to_latin(self, text):
        """
        Заменяет кириллические символы Н и В на латинские H и B.
        """
        translation_table = str.maketrans({'Н': 'H', 'В': 'B'})
        return text.translate(translation_table)

    def extract_prefix(self, file_name):
        """
        Преобразует имя файла в формат HBXXX.XXXXXX, добавляя точки между частями.
        Преобразует кириллицу в латиницу перед проверкой.
        """
        file_name = self.replace_cyrillic_to_latin(file_name)
        match = re.match(r'^(HB\d{3})[\s.-]?(\d{6})', file_name)

        if match:
            part1 = match.group(1)  # HBXXX
            part2 = match.group(2)  # XXXXXX
            return f"{part1}.{part2}"
        
        return None

    def create_escd_dict(self, file_path):
        """
        Читает xlsx файл и создает словарь из данных.
        Ключом является номер (например, 36), значением — название.
        """
        escd_dict = {}
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1]:
                escd_dict[str(row[0])] = row[1]  #  {'36': 'Суда, судовое оборудование'}
        
        return escd_dict
    

    # def move_files_to_folders(self, destination_folder):
    #     """
    #     Распределение файлов по папкам на основе префикса, системы ЕСКД РФ и обработки неправильных форматов.
    #     """
    #     file_names = self.get_file_names()
    #     seen = {}
    #     escd_dict = self.create_escd_dict("ESCD.xlsx")
    #     for file_name in file_names:
    #         prefix = self.extract_prefix(file_name)
    #         lat_file_name = self.replace_cyrillic_to_latin(file_name)

    #         if prefix:
    #             project_code = prefix[:5]  
    #             project_number = prefix[6:]  #

    #             if re.match(r'^HB\d{3}\.\d{6}', prefix):
    #                 target_folder = os.path.join(destination_folder, project_code)  

    
    #                 for i in range(2, len(project_number) + 1):
    #                     sub_folder = project_number[:i]  
                        
    #                     if sub_folder in escd_dict:
    #                         folder_name = f"{sub_folder} - {escd_dict[sub_folder]}" 
    #                     else:
    #                         folder_name = sub_folder 
    #                     target_folder = os.path.join(target_folder, folder_name)
    #             else:
    #                 target_folder = os.path.join(destination_folder, project_code, "Прочие файлы")
            
    #         elif re.match(r'^HB\d{3}-\d{3}', lat_file_name):
    #             project_code = lat_file_name[:5]  
    #             target_folder = os.path.join(destination_folder, project_code, "ОССЗ.Письма проектировщика")
            
    #         elif re.match(r'^120-\d{3}', lat_file_name):
    #             target_folder = os.path.join(destination_folder, "ОССЗ", "Письма")
            
    #         else:
    #             target_folder = os.path.join(destination_folder, "Прочие документы")
            
    #         extension = os.path.splitext(file_name)[1][1:].lower()
    #         if extension:
    #             target_folder = os.path.join(target_folder, extension)

    #         if not os.path.exists(target_folder):
    #             os.makedirs(target_folder)

    #         src_path = os.path.join(self.folder_path, file_name)
    #         dest_path = os.path.join(target_folder, file_name)

    #         if file_name not in seen:
    #             shutil.copy(src_path, dest_path)  
    #             seen[file_name] = True
    def move_files_to_folders(self, destination_folder):
        file_names = self.get_file_names()
        escd_dict = self.create_escd_dict("ESCD.xlsx")
        seen = {}

        for file_name in file_names:
            lat_file_name = self.replace_cyrillic_to_latin(file_name)
            prefix = self.extract_prefix(file_name)
            
            if prefix:
                target_folder = self.handle_prefix_case(destination_folder, prefix, escd_dict)
            elif re.match(r'^HB\d{3}-\d{3}', lat_file_name):
                target_folder = self.handle_specific_format_case(destination_folder, lat_file_name, "ОССЗ.Письма проектировщика")
                if re.match(r'^120-\d{3}', lat_file_name):
                    target_folder = self.handle_specific_format_case(destination_folder, lat_file_name, "ОССЗ.Письма")
            else:
                target_folder = os.path.join(destination_folder, "Прочие документы")

            target_folder = self.append_extension_folder(target_folder, file_name)
            self.copy_file_to_folder(file_name, target_folder, seen)

    def handle_prefix_case(self, destination_folder, prefix, escd_dict):
        project_code = prefix[:5]
        project_number = prefix[6:]
        target_folder = os.path.join(destination_folder, project_code)

        for i in range(2, len(project_number) + 1):
            sub_folder = project_number[:i]
            folder_name = f"{sub_folder} - {escd_dict.get(sub_folder, sub_folder)}"
            target_folder = os.path.join(target_folder, folder_name)

        return target_folder

    def handle_specific_format_case(self, destination_folder, file_name, specific_folder):
        project_code = file_name[:5]
        return os.path.join(destination_folder, project_code, specific_folder)

    def append_extension_folder(self, target_folder, file_name):
        extension = os.path.splitext(file_name)[1][1:].lower()
        if extension:
            target_folder = os.path.join(target_folder, extension)
        return target_folder

    def copy_file_to_folder(self, file_name, target_folder, seen):
        if not os.path.exists(target_folder):
            os.makedirs(target_folder)

        src_path = os.path.join(self.folder_path, file_name)
        dest_path = os.path.join(target_folder, file_name)

        if file_name not in seen:
            shutil.copy(src_path, dest_path)
            seen[file_name] = True

        
